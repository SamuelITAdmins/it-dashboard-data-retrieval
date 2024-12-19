import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

# Step 1: Fetch data from Google
url = "https://www.google.com"
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
}
response = requests.get(url, headers=headers)

# Check for successful response
if response.status_code == 200:
    # Parse the HTML using BeautifulSoup
    soup = BeautifulSoup(response.text, 'html.parser')
    # Extracting an example element (e.g., the Google homepage title and a few links)
    page_title = soup.title.string
    links = [a.get('href') for a in soup.find_all('a', href=True)][:5]  # First 5 links
    # print(soup.prettify())

    # Create an Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Google Data"

    # Write data to specific cells
    ws["A1"] = "Page Title"
    ws["B1"] = page_title  # Title goes into B1
    ws["A2"] = "Top Links"
    for i, link in enumerate(links, start=2):
        ws[f"B{i}"] = link  # Links go into column B, rows 2 to 6

    # Save the Excel file
    file_name = "Google_Data.xlsx"
    wb.save(file_name)
    print(f"Data saved to {file_name}")
else:
    print(f"Failed to retrieve data from {url}. HTTP Status Code: {response.status_code}")
