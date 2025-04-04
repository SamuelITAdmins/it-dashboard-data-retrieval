from openpyxl import Workbook
from merakidata import getMerakiData
from freshservicedata import getFreshServiceData
from azuredata import getAADUsers

# Create an Excel workbook
wb = Workbook()
switches_ws = wb.active
switches_ws.title = "Meraki Switches"

# Get Meraki Data
switch_downtimes, ap_downtimes = getMerakiData()

# Write Meraki Data
switches_ws["A1"] = "Switches"
switches_ws["B1"] = "Location"
switches_ws["C1"] = "Uptime"
switches_ws["D1"] = "Status"
for i, switch_downtime in enumerate(switch_downtimes, start=2):
    switches_ws[f"A{i}"] = switch_downtime['device_name']
    switches_ws[f"B{i}"] = switch_downtime['location']
    switches_ws[f"C{i}"] = switch_downtime['uptime_percentage']
    switches_ws[f"D{i}"] = switch_downtime['status']

ap_ws = wb.create_sheet(title="Meraki AP")

ap_ws["A1"] = "Access Points"
ap_ws["B1"] = "Location"
ap_ws["C1"] = "Uptime"
ap_ws["D1"] = "Status"
for i, ap_downtime in enumerate(ap_downtimes, start=2):
    ap_ws[f"A{i}"] = ap_downtime['device_name']
    ap_ws[f"B{i}"] = ap_downtime['location']
    ap_ws[f"C{i}"] = ap_downtime['uptime_percentage']
    ap_ws[f"D{i}"] = ap_downtime['status']

# Get FreshService Data
fs_data = getFreshServiceData()

# Write FreshService Data
fs_ws = wb.create_sheet(title="Freshservice")

fs_ws["A1"] = "Total Tickets (Last 7 Days)"
fs_ws["B1"] = "Unresolved Tickets (Last 7 Days)"
fs_ws["C1"] = "Resolved Tickets (Last 7 Days)"
fs_ws["D1"] = "SLA Resolution Compliance"
fs_ws["A2"] = fs_data['total_tickets']
fs_ws["B2"] = fs_data['unresolved_tickets']
fs_ws["C2"] = fs_data['resolved_tickets']
fs_ws["D2"] = fs_data['resolution_percentage']

# Get Azure active directory data
companyname = "Samuel Engineering"
users = getAADUsers(companyname)

# Create a new sheet for AAD user data
aad_ws = wb.create_sheet(title= "AAD Users")

# Header row
aad_ws["A1"] = "User Principal Name"
aad_ws["B1"] = "Job Title"
aad_ws["C1"] = "Department"
aad_ws["D1"] = "City"

# Fill rows with user data
for idx, user in enumerate(users, start=2):
    aad_ws[f"A{idx}"] = user.get("userPrincipalName", "")
    aad_ws[f"B{idx}"] = user.get("jobTitle", "")
    aad_ws[f"C{idx}"] = user.get("department", "")
    aad_ws[f"D{idx}"] = user.get("city", "")



# Save the Excel file
file_name = "IT Metric Dashboard Spreadsheet.xlsx"
wb.save(file_name)
print(f" Azure AD user data for '{companyname}' added to sheet '{aad_ws}' in '{file_name}'")
print(f"Data saved to {file_name}") 

